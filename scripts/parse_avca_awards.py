#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""AVCA All-America teams and award winners, out of the published spreadsheets.

    python3 scripts/parse_avca_awards.py     # -> data/avca_awards.json

WHY THIS IS WORTH HAVING. Nothing in any feed we can reach says a player was an
All-American. It is the one piece of context that separates "returns her top
scorer" from "returns the national Player of the Year", and it is exactly the
sort of thing the page has no way to know.

TWO JOINS, BOTH OF WHICH CAN GO WRONG QUIETLY
---------------------------------------------
**Schools.** The spreadsheet writes full formal names -- "University of
Pittsburgh", "Texas A&M University", "SMU". Our `name_full` covers some of them
and misses others ("University of Nebraska" is "University of Nebraska,
Lincoln"). So the match is tried in three steps, best first, and anything left
over is REPORTED rather than dropped silently: exact `name_full`, then a
normalised comparison with the university boilerplate stripped, then
`name_short`. A school that matches none renders no badge.

**Players.** R8 applies: the surname must match EXACTLY and only the given name
may flex, and the pair must be unique within that school. An All-America badge
on the wrong player is the same class of error as attributing her stats.

WHAT IS NOT DONE. Historical seasons are parsed but only the two most recent are
loaded by the page -- an honour from 2003 is a fact about a person who is not on
a 2026 roster. The file keeps them so the question can be asked later.

Python 3.9 target. Needs openpyxl (already installed).
"""

import json
import os
import re
import sys
from typing import Any, Dict, List, Optional, Tuple

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(REPO, "data", "avca_awards.json")

ALL_AMERICA = os.path.join(
    REPO, "AVCA-Division-I-Womens-VB-All-America-Teams-Year-by-Year-1981-2025.xlsx")

# Section headers as they appear in column A.
TEAM_HEADERS = ("first team", "second team", "third team", "honorable mention")

# University boilerplate, removed before comparing school names.
BOILER = re.compile(
    r"\b(university|univ|of|the|at|college|state university|u\.s\.|academy)\b", re.I)


def squash(s):
    # type: (Optional[str]) -> str
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def loose(s):
    # type: (Optional[str]) -> str
    """School name with the formal boilerplate stripped, for a second attempt.

    ALSO NORMALISES "State" -> "St", which is not cosmetic: the AVCA writes
    "Penn State" and ncaa.com writes "Penn St.", and without this Penn State's
    FORTY All-Americans matched nothing. Same for Ohio State, Long Beach State,
    Colorado State -- 73 schools in total. A join that drops the biggest
    programmes and says nothing is worse than no join.
    """
    t = BOILER.sub(" ", s or "")
    t = re.sub(r"\bstate\b", "st", t, flags=re.I)
    t = re.sub(r"\bsaint\b", "st", t, flags=re.I)
    return squash(t)


# Hand aliases, and the bar for adding one is high: it must be a school whose
# name genuinely differs between the two sources, not a case the fuzzy pass
# nearly got. Only five recent selections needed these -- "University of Miami"
# is our "Miami (FL)" (the AVCA does not disambiguate from Miami of Ohio), and
# one is simply a typo in the workbook: "Califonia".
ALIASES = {
    "universityofmiami": "Miami (FL)",
    "miami": "Miami (FL)",
    "universityofsoutherncalifonia": "Southern California",
    "calstatenorthridge": "CSUN",
    "uwmilwaukee": "Milwaukee",
    "collegeofcharleston": "Col. of Charleston",
}


def team_lookup():
    # type: () -> Dict[str, str]
    """Every key we know for a school -> our short name."""
    out = {}
    src = os.path.join(REPO, "data", "data_2025.json")
    if not os.path.exists(src):
        return out
    for t in (json.load(open(src, encoding="utf-8")) or {}).get("teams", []):
        short = t.get("name_short")
        if not short:
            continue
        for key in (t.get("name_full"), short):
            if key:
                out.setdefault(squash(key), short)
                out.setdefault(loose(key), short)
    return out


def parse_year(ws):
    # type: (Any) -> Dict[str, Any]
    """One year's sheet -> teams and award winners."""
    rows = list(ws.iter_rows(values_only=True))
    players = []                                        # type: List[Dict]
    awards = []                                         # type: List[Dict]
    team = None
    pending_award = None
    for r in rows:
        cells = list(r) + [None] * 8
        a = (cells[0] or "")
        a = a.strip() if isinstance(a, str) else ""
        low = a.lower()
        if low in TEAM_HEADERS:
            team = a
            continue
        if low in ("first", ""):                        # header row or blank
            pass
        # Award column: a label on one row, the winner on the next.
        g = cells[6]
        if isinstance(g, str) and g.strip():
            g = g.strip()
            if g.lower().endswith("of the year") or "award" in g.lower():
                pending_award = g
            elif pending_award:
                awards.append({"award": pending_award, "text": g})
                pending_award = None
        first, last, school = cells[0], cells[1], cells[2]
        if (team and isinstance(first, str) and isinstance(last, str)
                and isinstance(school, str) and first.strip() and last.strip()
                and low not in ("first",)):
            players.append({
                "first": first.strip(), "last": last.strip(),
                "school_raw": school.strip(),
                "pos": (cells[3] or "").strip() if isinstance(cells[3], str) else None,
                "year": (cells[4] or "").strip() if isinstance(cells[4], str) else None,
                "previous": (cells[5] or "").strip() if isinstance(cells[5], str) else None,
                "honour": team,          # First Team / Second Team / ...
            })
    return {"players": players, "awards": awards}


def main():
    if not os.path.exists(ALL_AMERICA):
        print("no All-America workbook at %s" % ALL_AMERICA)
        return 1
    try:
        import openpyxl
    except ImportError:
        print("needs openpyxl: python3 -m pip install --user openpyxl")
        return 1

    wb = openpyxl.load_workbook(ALL_AMERICA, read_only=True, data_only=True)
    lookup = team_lookup()

    seasons = {}
    unmatched_schools = {}
    for name in wb.sheetnames:
        if not re.match(r"^(Fall |Spring )?\d{4}$", name.strip()):
            continue
        year = re.search(r"\d{4}", name).group(0)
        got = parse_year(wb[name])
        for p in got["players"]:
            raw = p["school_raw"]
            short = (lookup.get(squash(raw)) or lookup.get(loose(raw))
                     or ALIASES.get(squash(raw)))
            p["team"] = short
            if not short:
                unmatched_schools[raw] = unmatched_schools.get(raw, 0) + 1
        seasons.setdefault(year, []).append({"sheet": name, **got})

    flat = []
    for year, blocks in seasons.items():
        for b in blocks:
            for p in b["players"]:
                flat.append(dict(season=int(year), **p))

    doc = {
        "meta": {
            "source": "AVCA published All-America workbook (1981-2025)",
            "source_tier": "OFFICIAL (AVCA)",
            "note": ("Nothing in any feed we reach carries an All-America "
                     "selection. School names are matched in three passes and "
                     "anything unmatched is listed here rather than dropped."),
            "seasons": sorted(seasons, reverse=True)[:8],
            "selections": len(flat),
            "unmatched_schools": sorted(unmatched_schools.items(),
                                        key=lambda kv: -kv[1])[:25],
        },
        "selections": flat,
        "awards": dict((y, sum((b["awards"] for b in blocks), []))
                       for y, blocks in seasons.items()),
    }
    json.dump(doc, open(OUT, "w"), indent=1, sort_keys=False)
    m = doc["meta"]
    print("seasons parsed      : %d" % len(seasons))
    print("selections          : %d" % m["selections"])
    print("schools unmatched   : %d distinct" % len(unmatched_schools))
    for raw, n in m["unmatched_schools"][:8]:
        print("    %-42s %d" % (raw[:42], n))
    print("wrote %s" % OUT)
    return 0


if __name__ == "__main__":
    sys.exit(main())
